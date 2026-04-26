# =============================================================================
# metrics.py — Generate hbfa_metrics.xlsx from coverage + crash CSVs
# =============================================================================

"""Combine per-harness coverage and crash data into a single Excel workbook."""

import sys
from pathlib import Path

from .config import METRICS_XLSX

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl",
          file=sys.stderr)
    sys.exit(1)


HEADERS = [
    "Harness",
    "edk2 Component",
    "Lines Hit",
    "Lines Total",
    "Line Coverage %",
    "Functions Hit",
    "Functions Total",
    "Function Coverage %",
    "Crashes (Raw)",
    "Crashes (Unique)",
    "Error Types",
]


def _style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="305496")
    cell.alignment = Alignment(horizontal="center", vertical="center")


# Two pale fills used to alternate row groups (one harness = one group),
# plus a slightly darker blue used for the ALL aggregate row inside a group.
_GROUP_FILLS    = (PatternFill("solid", fgColor="FFFFFF"),
                   PatternFill("solid", fgColor="EAF2FB"))
_AGG_FILL       = PatternFill("solid", fgColor="D9E5F4")
_THIN           = Side(style="thin", color="BFBFBF")
_THICK          = Side(style="medium", color="305496")


def write_metrics(coverage_rows: list, crash_rows: list,
                  out_path: Path = METRICS_XLSX) -> Path:
    # coverage_rows now has one entry per (harness, component) pair, so
    # group them by harness to emit one Excel row per pair. Crash data
    # is per-harness; replicate it across the harness's component rows.
    cov_by_name: dict = {}
    for r in coverage_rows:
        cov_by_name.setdefault(r["harness"], []).append(r)
    crash_by_name = {r["harness"]: r for r in crash_rows}
    names = sorted(set(cov_by_name) | set(crash_by_name))

    wb = Workbook()
    ws = wb.active
    ws.title = "Metrics"

    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        _style_header(c)

    r_idx = 2
    n_cols = len(HEADERS)
    for grp_idx, name in enumerate(names):
        cr = crash_by_name.get(name, {})
        cov_list = cov_by_name.get(name) or [{}]
        group_fill = _GROUP_FILLS[grp_idx % 2]
        first_row = r_idx
        for cov in cov_list:
            component = (cov.get("component")
                         or cov.get("edk2_source")
                         or cr.get("edk2_source", ""))
            is_agg = isinstance(component, str) and component.startswith("ALL")
            row = [
                name,
                component,
                cov.get("lines_hit", 0),
                cov.get("lines_total", 0),
                cov.get("line_coverage_pct", 0.0),
                cov.get("functions_hit", 0),
                cov.get("functions_total", 0),
                cov.get("function_coverage_pct", 0.0),
                cr.get("crashes_total", 0),
                cr.get("crashes_unique", 0),
                cr.get("error_types", ""),
            ]
            fill = _AGG_FILL if is_agg else group_fill
            for i, v in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=i, value=v)
                cell.fill = fill
                cell.border = Border(left=_THIN, right=_THIN,
                                     top=_THIN, bottom=_THIN)
                if is_agg:
                    cell.font = Font(bold=True)
                if i in (3, 4, 6, 7, 9, 10):
                    cell.alignment = Alignment(horizontal="right")
                if i == 5 or i == 8:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "0.00"
            r_idx += 1
        last_row = r_idx - 1
        # Merge the harness column across the group's rows for a clean
        # one-row-per-harness visual grouping; keep top-aligned.
        if last_row > first_row:
            ws.merge_cells(start_row=first_row, end_row=last_row,
                           start_column=1, end_column=1)
        top_cell = ws.cell(row=first_row, column=1)
        top_cell.alignment = Alignment(horizontal="left", vertical="center")
        top_cell.font = Font(bold=True)
        # Thicker bottom border on the last row of each group to visually
        # separate harnesses from each other.
        for col in range(1, n_cols + 1):
            c = ws.cell(row=last_row, column=col)
            existing = c.border
            c.border = Border(left=existing.left, right=existing.right,
                              top=existing.top, bottom=_THICK)

    # Column widths
    widths = [34, 56, 11, 12, 16, 14, 16, 20, 14, 16, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    _write_component_summary(wb, coverage_rows)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[metrics] wrote {out_path}")
    return out_path


# --------------------------------------------------------------------------
# Second sheet: unique edk2 components with averaged coverage across every
# harness that targets them. Rows hit by a single harness still appear so
# the sheet is a complete picture of component-level reach.
# --------------------------------------------------------------------------

_COMPONENT_HEADERS = [
    "edk2 Component",
    "Harness Count",
    "Harnesses",
    "Union Lines Hit",
    "Lines Total",
    "Union Line Coverage %",
    "Union Functions Hit",
    "Functions Total",
    "Union Function Coverage %",
]


def _write_component_summary(wb, coverage_rows: list) -> None:
    ws = wb.create_sheet("By Component")

    for i, h in enumerate(_COMPONENT_HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        _style_header(c)

    # Aggregate per component by *unioning* the per-harness hit sets so
    # a line/function covered by ANY harness counts once at the
    # component level. Synthetic "ALL (...)" rows are skipped to avoid
    # double-counting multi-component harnesses.
    by_comp: dict = {}
    for r in coverage_rows:
        comp = r.get("component") or ""
        if not comp or comp.startswith("ALL"):
            continue
        b = by_comp.setdefault(comp, {
            "harnesses":     set(),
            "line_keys":     set(),
            "fn_keys":       set(),
            "lines_total":   0,
            "funcs_total":   0,
        })
        b["harnesses"].add(r["harness"])
        b["line_keys"] |= r.get("_lines_hit_keys") or set()
        b["fn_keys"]   |= r.get("_fns_hit_keys")   or set()
        # Component totals are deterministic per component, but a row
        # might have 0 if lcov failed for that harness — keep the max.
        b["lines_total"] = max(b["lines_total"], int(r.get("lines_total") or 0))
        b["funcs_total"] = max(b["funcs_total"], int(r.get("functions_total") or 0))

    r_idx = 2
    for comp in sorted(by_comp):
        b = by_comp[comp]
        u_lines = len(b["line_keys"])
        u_funcs = len(b["fn_keys"])
        # Cap at the on-disk total in case stale .info entries leak
        # extra hits.
        u_lines = min(u_lines, b["lines_total"]) if b["lines_total"] else u_lines
        u_funcs = min(u_funcs, b["funcs_total"]) if b["funcs_total"] else u_funcs
        line_pct = round(100.0 * u_lines / b["lines_total"], 2) if b["lines_total"] else 0.0
        fn_pct   = round(100.0 * u_funcs / b["funcs_total"], 2) if b["funcs_total"] else 0.0
        row = [
            comp,
            len(b["harnesses"]),
            ", ".join(sorted(b["harnesses"])),
            u_lines, b["lines_total"], line_pct,
            u_funcs, b["funcs_total"], fn_pct,
        ]
        fill = _GROUP_FILLS[(r_idx - 2) % 2]
        for i, v in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=i, value=v)
            cell.fill = fill
            cell.border = Border(left=_THIN, right=_THIN,
                                 top=_THIN, bottom=_THIN)
            if i in (6, 9):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "0.00"
            elif i in (2, 4, 5, 7, 8):
                cell.alignment = Alignment(horizontal="right")
        r_idx += 1

    widths = [56, 14, 60, 16, 14, 22, 20, 16, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if r_idx > 2:
        ws.auto_filter.ref = ws.dimensions
