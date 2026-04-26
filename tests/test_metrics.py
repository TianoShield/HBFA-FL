"""Tests for InsightsExtractor.metrics (Excel writer)."""
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from InsightsExtractor import metrics  # noqa: E402

try:
    import openpyxl
except ImportError:                         # pragma: no cover
    openpyxl = None


def _cov_row(harness, comp, lines_hit_keys=None, fns_hit_keys=None,
             lines_total=10, functions_total=4):
    lh = len(lines_hit_keys or set())
    fh = len(fns_hit_keys or set())
    return {
        "harness": harness,
        "component": comp,
        "lines_total": lines_total,
        "lines_hit": lh,
        "line_coverage_pct": round(100.0 * lh / lines_total, 2)
        if lines_total else 0.0,
        "functions_total": functions_total,
        "functions_hit": fh,
        "function_coverage_pct": round(100.0 * fh / functions_total, 2)
        if functions_total else 0.0,
        "_lines_hit_keys": set(lines_hit_keys or set()),
        "_fns_hit_keys": set(fns_hit_keys or set()),
    }


@unittest.skipIf(openpyxl is None, "openpyxl not installed")
class WriteMetricsTest(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        self._tmp.close()
        self.path = Path(self._tmp.name)

    def tearDown(self):
        self.path.unlink(missing_ok=True)

    def test_two_sheets_with_headers(self):
        cov = [_cov_row("A", "Pkg/X")]
        out = metrics.write_metrics(cov, [], self.path)
        wb = openpyxl.load_workbook(out)
        self.assertEqual(wb.sheetnames, ["Metrics", "By Component"])
        self.assertEqual([c.value for c in wb["Metrics"][1]], metrics.HEADERS)
        self.assertEqual([c.value for c in wb["By Component"][1]],
                         metrics._COMPONENT_HEADERS)

    def test_union_math_disjoint_and_overlap(self):
        cov = [
            _cov_row("A", "Pkg/X",
                     lines_hit_keys={"f.c:1", "f.c:2", "f.c:3"},
                     fns_hit_keys={"f.c::a", "f.c::b"},
                     lines_total=10, functions_total=4),
            _cov_row("B", "Pkg/X",
                     lines_hit_keys={"f.c:3", "f.c:4", "f.c:5"},
                     fns_hit_keys={"f.c::b", "f.c::c"},
                     lines_total=10, functions_total=4),
        ]
        metrics.write_metrics(cov, [], self.path)
        wb = openpyxl.load_workbook(self.path)
        ws = wb["By Component"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        comp, hcount, hlist, ulines, ltotal, lpct, ufns, ftotal, fpct = rows[0]
        self.assertEqual(comp, "Pkg/X")
        self.assertEqual(hcount, 2)
        self.assertEqual(hlist, "A, B")
        self.assertEqual(ulines, 5)        # union {1,2,3,4,5}
        self.assertEqual(ltotal, 10)
        self.assertEqual(lpct, 50.0)
        self.assertEqual(ufns, 3)          # union {a,b,c}
        self.assertEqual(ftotal, 4)
        self.assertEqual(fpct, 75.0)

    def test_union_caps_hits_at_total(self):
        # If lcov reports stale/extra lines, they must not push union > total.
        cov = [_cov_row("A", "Pkg/Y",
                        lines_hit_keys={f"f.c:{i}" for i in range(20)},
                        lines_total=10)]
        metrics.write_metrics(cov, [], self.path)
        wb = openpyxl.load_workbook(self.path)
        ulines = wb["By Component"].cell(row=2, column=4).value
        ltotal = wb["By Component"].cell(row=2, column=5).value
        self.assertLessEqual(ulines, ltotal)
        self.assertEqual(ulines, 10)

    def test_all_aggregate_rows_excluded_from_summary(self):
        cov = [
            _cov_row("A", "Pkg/X",
                     lines_hit_keys={"f.c:1"}, fns_hit_keys={"f.c::a"}),
            _cov_row("A", "Pkg/Y",
                     lines_hit_keys={"g.c:1"}, fns_hit_keys={"g.c::g"}),
            _cov_row("A", "ALL (Pkg/X;Pkg/Y)",
                     lines_hit_keys={"f.c:1", "g.c:1"},
                     fns_hit_keys={"f.c::a", "g.c::g"}),
        ]
        metrics.write_metrics(cov, [], self.path)
        wb = openpyxl.load_workbook(self.path)
        comps = [r[0] for r in wb["By Component"].iter_rows(min_row=2,
                                                            values_only=True)]
        self.assertCountEqual(comps, ["Pkg/X", "Pkg/Y"])

    def test_metrics_sheet_one_row_per_component(self):
        cov = [
            _cov_row("A", "Pkg/X", lines_hit_keys={"x.c:1"}),
            _cov_row("A", "Pkg/Y", lines_hit_keys={"y.c:1"}),
            _cov_row("B", "Pkg/X", lines_hit_keys={"x.c:2"}),
        ]
        crashes = [{"harness": "A", "edk2_source": "Pkg/X",
                    "crashes_total": 7, "crashes_unique": 2,
                    "error_types": "SEGV"}]
        metrics.write_metrics(cov, crashes, self.path)
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Metrics"]
        # 1 header row + 3 data rows
        self.assertEqual(ws.max_row, 4)


if __name__ == "__main__":
    unittest.main()
